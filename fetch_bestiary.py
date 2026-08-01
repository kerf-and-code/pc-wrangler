"""
fetch_bestiary.py

Pulls every creature stat block from the Open5e v2 API and writes them to a
single .xlsx workbook.

Open5e only serves openly licensed material (OGL / CC-BY), including:
  srd-2014  System Reference Document 5.1  (WotC, CC-BY-4.0)
  srd-2024  System Reference Document 5.2  (WotC, CC-BY-4.0)
  tob, tob-2023, tob2, tob3, ccdx, bfrd    (Kobold Press, OGL)
  a5e-mm                                   (EN Publishing, OGL)
  and others

Usage (Windows, from anywhere):
    py -m pip install openpyxl
    py fetch_bestiary.py
    py fetch_bestiary.py --documents srd-2014,srd-2024
    py fetch_bestiary.py --gamesystem 5e-2024 --out srd52.xlsx
    py fetch_bestiary.py --list-documents
    py fetch_bestiary.py --cache creatures.json

Sheets written:
    Monsters      one row per creature, wide stat-block columns
    Actions       one row per action  (long format, for DB import)
    Traits        one row per trait   (long format, for DB import)
    Sources       one row per source document, with counts and license links
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.request
from typing import Any, Dict, Iterable, List, Optional

API_ROOT = "https://api.open5e.com/v2/creatures/?limit=500"
USER_AGENT = "bestiary-export/1.0 (personal spreadsheet build)"
CELL_LIMIT = 32000  # Excel hard cap is 32767; leave headroom for the ellipsis

ABILITIES = ["strength", "dexterity", "constitution", "intelligence", "wisdom", "charisma"]
ABBREV = {
    "strength": "STR",
    "dexterity": "DEX",
    "constitution": "CON",
    "intelligence": "INT",
    "wisdom": "WIS",
    "charisma": "CHA",
}
SPEED_KEYS = ["walk", "fly", "swim", "climb", "burrow"]


# ---------------------------------------------------------------- fetching


def fetch_page(url: str, retries: int = 4) -> Dict[str, Any]:
    """GET one page of JSON, with a simple backoff on transient failures."""
    last_error: Optional[Exception] = None
    for attempt in range(retries):
        try:
            request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(request, timeout=60) as response:
                return json.loads(response.read().decode("utf-8"))
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as error:
            last_error = error
            wait = 2 ** attempt
            print(f"  retry {attempt + 1}/{retries} in {wait}s ({error})", file=sys.stderr)
            time.sleep(wait)
    raise RuntimeError(f"giving up on {url}: {last_error}")


def fetch_all_creatures() -> List[Dict[str, Any]]:
    """Walk the paginated creature list by following the API's own next link."""
    creatures: List[Dict[str, Any]] = []
    url: Optional[str] = API_ROOT
    page = 0
    while url:
        page += 1
        payload = fetch_page(url)
        batch = payload.get("results", [])
        creatures.extend(batch)
        total = payload.get("count")
        print(f"  page {page}: +{len(batch)} (have {len(creatures)} of {total})")
        url = payload.get("next")
    return creatures


def load_or_fetch(cache_path: Optional[str]) -> List[Dict[str, Any]]:
    if cache_path and os.path.exists(cache_path):
        print(f"Reading cached creatures from {cache_path}")
        with open(cache_path, "r", encoding="utf-8") as handle:
            return json.load(handle)

    print("Fetching creatures from Open5e ...")
    creatures = fetch_all_creatures()

    if cache_path:
        with open(cache_path, "w", encoding="utf-8") as handle:
            json.dump(creatures, handle)
        print(f"Cached raw JSON to {cache_path}")
    return creatures


# ---------------------------------------------------------------- shaping


def clean(value: Any) -> str:
    """Excel refuses most control characters, so strip them out."""
    if value is None:
        return ""
    text = str(value)
    text = "".join(char for char in text if char >= " " or char in "\n\t")
    if len(text) > CELL_LIMIT:
        text = text[:CELL_LIMIT] + " ...[truncated]"
    return text


def nested_name(node: Any) -> str:
    if isinstance(node, dict):
        return clean(node.get("name", ""))
    return clean(node)


def cr_display(value: Any) -> str:
    """0.25 reads as 1/4 on a stat block, so print it that way."""
    if value is None:
        return ""
    fractions = {0.125: "1/8", 0.25: "1/4", 0.5: "1/2"}
    try:
        number = float(value)
    except (TypeError, ValueError):
        return clean(value)
    if number in fractions:
        return fractions[number]
    if number.is_integer():
        return str(int(number))
    return str(number)


def signed(value: Any) -> str:
    if value is None:
        return ""
    try:
        number = int(value)
    except (TypeError, ValueError):
        return clean(value)
    return f"+{number}" if number >= 0 else str(number)


def bonus_map(mapping: Any) -> str:
    """Render {'stealth': 6, 'perception': 3} as 'Perception +3, Stealth +6'."""
    if not isinstance(mapping, dict) or not mapping:
        return ""
    parts = []
    for key in sorted(mapping):
        label = key.replace("_", " ").title()
        parts.append(f"{label} {signed(mapping[key])}")
    return ", ".join(parts)


def saves_display(creature: Dict[str, Any]) -> str:
    """Only the proficient saves live in saving_throws; the _all copy is noise."""
    saves = creature.get("saving_throws")
    if not isinstance(saves, dict) or not saves:
        return ""
    parts = []
    for ability in ABILITIES:
        if ability in saves:
            parts.append(f"{ABBREV[ability]} {signed(saves[ability])}")
    return ", ".join(parts)


def senses_display(creature: Dict[str, Any]) -> str:
    parts = []
    for field, label in (
        ("blindsight_range", "blindsight"),
        ("darkvision_range", "darkvision"),
        ("tremorsense_range", "tremorsense"),
        ("truesight_range", "truesight"),
    ):
        value = creature.get(field)
        if value:
            parts.append(f"{label} {value} ft.")
    passive = creature.get("passive_perception")
    if passive is not None:
        parts.append(f"passive Perception {passive}")
    return ", ".join(parts)


def entries_text(entries: Iterable[Dict[str, Any]]) -> str:
    """Collapse a list of {name, desc} into one readable cell."""
    chunks = []
    for entry in entries:
        name = clean(entry.get("name", ""))
        desc = clean(entry.get("desc", ""))
        chunks.append(f"{name}. {desc}".strip())
    return "\n\n".join(chunk for chunk in chunks if chunk)


def usage_display(entry: Dict[str, Any]) -> str:
    limits = entry.get("usage_limits")
    if not isinstance(limits, dict) or not limits:
        return ""
    kind = clean(limits.get("type", "")).replace("_", " ").lower()
    param = limits.get("param")
    return f"{kind} {param}".strip()


def sorted_actions(creature: Dict[str, Any]) -> List[Dict[str, Any]]:
    actions = creature.get("actions") or []
    return sorted(actions, key=lambda entry: (entry.get("order_in_statblock") or 0, entry.get("name") or ""))


def actions_of_type(creature: Dict[str, Any], wanted: str) -> List[Dict[str, Any]]:
    return [a for a in sorted_actions(creature) if (a.get("action_type") or "ACTION") == wanted]


def other_action_types(creature: Dict[str, Any]) -> List[Dict[str, Any]]:
    known = {"ACTION", "BONUS_ACTION", "REACTION", "LEGENDARY_ACTION"}
    return [a for a in sorted_actions(creature) if (a.get("action_type") or "ACTION") not in known]


MONSTER_COLUMNS = [
    "key",
    "name",
    "source",
    "source_key",
    "publisher",
    "game_system",
    "cr",
    "xp",
    "size",
    "type",
    "alignment",
    "ac",
    "armor_detail",
    "hp",
    "hit_dice",
    "initiative",
    "speed_walk",
    "speed_fly",
    "speed_swim",
    "speed_climb",
    "speed_burrow",
    "STR",
    "DEX",
    "CON",
    "INT",
    "WIS",
    "CHA",
    "STR_mod",
    "DEX_mod",
    "CON_mod",
    "INT_mod",
    "WIS_mod",
    "CHA_mod",
    "saving_throws",
    "skills",
    "damage_vulnerabilities",
    "damage_resistances",
    "damage_immunities",
    "condition_immunities",
    "senses",
    "languages",
    "environments",
    "trait_count",
    "action_count",
    "traits",
    "actions",
    "bonus_actions",
    "reactions",
    "legendary_actions",
    "other_entries",
    "source_url",
]


def monster_row(creature: Dict[str, Any]) -> List[Any]:
    document = creature.get("document") or {}
    scores = creature.get("ability_scores") or {}
    mods = creature.get("modifiers") or {}
    speeds = creature.get("speed_all") or creature.get("speed") or {}
    defenses = creature.get("resistances_and_immunities") or {}
    languages = creature.get("languages") or {}

    row: List[Any] = [
        clean(creature.get("key")),
        clean(creature.get("name")),
        nested_name(document),
        clean(document.get("key")),
        nested_name((document.get("publisher") or {})),
        nested_name((document.get("gamesystem") or {})),
        cr_display(creature.get("challenge_rating")),
        creature.get("experience_points"),
        nested_name(creature.get("size")),
        nested_name(creature.get("type")),
        clean(creature.get("alignment")),
        creature.get("armor_class"),
        clean(creature.get("armor_detail")),
        creature.get("hit_points"),
        clean(creature.get("hit_dice")),
        signed(creature.get("initiative_bonus")),
    ]

    for key in SPEED_KEYS:
        value = speeds.get(key) if isinstance(speeds, dict) else None
        row.append(value if value else None)

    for ability in ABILITIES:
        row.append(scores.get(ability))
    for ability in ABILITIES:
        row.append(signed(mods.get(ability)) if mods.get(ability) is not None else "")

    traits = creature.get("traits") or []
    actions = actions_of_type(creature, "ACTION")
    bonus = actions_of_type(creature, "BONUS_ACTION")
    reactions = actions_of_type(creature, "REACTION")
    legendary = actions_of_type(creature, "LEGENDARY_ACTION")
    other = other_action_types(creature)

    row.extend(
        [
            saves_display(creature),
            bonus_map(creature.get("skill_bonuses")),
            clean(defenses.get("damage_vulnerabilities_display")),
            clean(defenses.get("damage_resistances_display")),
            clean(defenses.get("damage_immunities_display")),
            clean(defenses.get("condition_immunities_display")),
            senses_display(creature),
            clean(languages.get("as_string") if isinstance(languages, dict) else languages),
            ", ".join(nested_name(env) for env in (creature.get("environments") or [])),
            len(traits),
            len(creature.get("actions") or []),
            entries_text(traits),
            entries_text(actions),
            entries_text(bonus),
            entries_text(reactions),
            entries_text(legendary),
            entries_text(other),
            clean(document.get("permalink")),
        ]
    )
    return row


ACTION_COLUMNS = [
    "creature_key",
    "creature_name",
    "source_key",
    "action_type",
    "order",
    "name",
    "desc",
    "usage",
    "legendary_cost",
    "to_hit",
    "reach_ft",
    "range_ft",
    "long_range_ft",
    "damage_dice",
    "damage_bonus",
    "damage_type",
]


def action_rows(creature: Dict[str, Any]) -> List[List[Any]]:
    document = creature.get("document") or {}
    rows: List[List[Any]] = []
    for entry in sorted_actions(creature):
        attacks = entry.get("attacks") or []
        attack = attacks[0] if attacks else {}
        dice_count = attack.get("damage_die_count")
        dice_type = attack.get("damage_die_type")
        dice = f"{dice_count}{clean(dice_type).lower()}" if dice_count and dice_type else ""
        damage_type = attack.get("extra_damage_type") or attack.get("damage_type")
        rows.append(
            [
                clean(creature.get("key")),
                clean(creature.get("name")),
                clean(document.get("key")),
                clean(entry.get("action_type") or "ACTION"),
                entry.get("order_in_statblock"),
                clean(entry.get("name")),
                clean(entry.get("desc")),
                usage_display(entry),
                entry.get("legendary_action_cost"),
                signed(attack.get("to_hit_mod")) if attack.get("to_hit_mod") is not None else "",
                attack.get("reach"),
                attack.get("range"),
                attack.get("long_range"),
                dice,
                attack.get("damage_bonus"),
                nested_name(damage_type),
            ]
        )
    return rows


TRAIT_COLUMNS = ["creature_key", "creature_name", "source_key", "name", "desc"]


def trait_rows(creature: Dict[str, Any]) -> List[List[Any]]:
    document = creature.get("document") or {}
    rows = []
    for entry in creature.get("traits") or []:
        rows.append(
            [
                clean(creature.get("key")),
                clean(creature.get("name")),
                clean(document.get("key")),
                clean(entry.get("name")),
                clean(entry.get("desc")),
            ]
        )
    return rows


SOURCE_COLUMNS = ["source_key", "source", "publisher", "game_system", "creatures", "permalink"]


def source_rows(creatures: List[Dict[str, Any]]) -> List[List[Any]]:
    seen: Dict[str, List[Any]] = {}
    for creature in creatures:
        document = creature.get("document") or {}
        key = clean(document.get("key"))
        if key not in seen:
            seen[key] = [
                key,
                nested_name(document),
                nested_name(document.get("publisher") or {}),
                nested_name(document.get("gamesystem") or {}),
                0,
                clean(document.get("permalink")),
            ]
        seen[key][4] += 1
    return sorted(seen.values(), key=lambda row: row[0])


# ---------------------------------------------------------------- writing


def write_workbook(creatures: List[Dict[str, Any]], out_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    body_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44546A")
    header_align = Alignment(horizontal="left", vertical="center")

    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(title: str, columns: List[str], rows: List[List[Any]], widths: Dict[str, int]) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(columns)
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
        sheet.freeze_panes = "C2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{sheet.max_row}"
        for index, name in enumerate(columns, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = widths.get(name, 14)
        print(f"  {title}: {len(rows)} rows")

    monster_widths = {
        "name": 30,
        "key": 28,
        "source": 30,
        "publisher": 20,
        "game_system": 18,
        "alignment": 18,
        "armor_detail": 22,
        "saving_throws": 24,
        "skills": 30,
        "damage_vulnerabilities": 22,
        "damage_resistances": 28,
        "damage_immunities": 28,
        "condition_immunities": 28,
        "senses": 32,
        "languages": 30,
        "environments": 30,
        "traits": 60,
        "actions": 60,
        "bonus_actions": 40,
        "reactions": 40,
        "legendary_actions": 40,
        "other_entries": 30,
        "source_url": 40,
    }

    add_sheet(
        "Monsters",
        MONSTER_COLUMNS,
        [monster_row(creature) for creature in creatures],
        monster_widths,
    )

    actions: List[List[Any]] = []
    traits: List[List[Any]] = []
    for creature in creatures:
        actions.extend(action_rows(creature))
        traits.extend(trait_rows(creature))

    add_sheet("Actions", ACTION_COLUMNS, actions, {"creature_name": 30, "name": 26, "desc": 90, "creature_key": 28})
    add_sheet("Traits", TRAIT_COLUMNS, traits, {"creature_name": 30, "name": 26, "desc": 90, "creature_key": 28})
    add_sheet(
        "Sources",
        SOURCE_COLUMNS,
        source_rows(creatures),
        {"source": 36, "publisher": 22, "game_system": 20, "permalink": 60},
    )

    workbook.save(out_path)


# ---------------------------------------------------------------- app json export
#
# The Six Axes app loads monsters from two per-edition files, lib/srd/monsters-2024.json and
# lib/srd/monsters-2014.json, each a flat array of monster objects. This exporter converts Open5e
# creatures into that exact shape and routes each creature into the file matching its game system,
# so third-party OGL content (mostly 5e-2014-compatible) lands in the 2014 file and SRD 5.2 content
# lands in the 2024 file. The shape here must stay in sync with statBlockFromMonster() in the app's
# lib/stat-blocks.ts and the flat monster fields the builder reads.


def _named_entries(entries: Iterable[Dict[str, Any]]) -> List[Dict[str, str]]:
    """Open5e {name, desc, ...} -> the app's {name, desc} pairs, in stat-block order."""
    out: List[Dict[str, str]] = []
    for entry in entries:
        name = clean(entry.get("name", "")).strip()
        desc = clean(entry.get("desc", "")).strip()
        if name or desc:
            out.append({"name": name, "desc": desc})
    return out


def _speed_string(creature: Dict[str, Any]) -> str:
    """Render the speed dict as the app's single speed string, e.g. '10 ft., Swim 40 ft.'."""
    speeds = creature.get("speed_all") or creature.get("speed") or {}
    if not isinstance(speeds, dict) or not speeds:
        return ""
    parts: List[str] = []
    walk = speeds.get("walk")
    if walk:
        parts.append(f"{walk} ft.")
    for key in ("burrow", "climb", "fly", "swim"):
        value = speeds.get(key)
        if value:
            label = key.capitalize()
            hover = " (hover)" if key == "fly" and speeds.get("hover") else ""
            parts.append(f"{label} {value} ft.{hover}")
    return ", ".join(parts)


def app_monster(creature: Dict[str, Any]) -> Dict[str, Any]:
    """Convert one Open5e creature into the app's flat monster object."""
    scores = creature.get("ability_scores") or {}
    document = creature.get("document") or {}
    languages = creature.get("languages") or {}
    defenses = creature.get("resistances_and_immunities") or {}

    def _score(ability: str) -> int:
        value = scores.get(ability)
        try:
            return int(value)
        except (TypeError, ValueError):
            return 10

    def _split(display: Any) -> List[str]:
        text = clean(display).strip()
        return [part.strip() for part in text.split(",") if part.strip()] if text else []

    return {
        "name": clean(creature.get("name")),
        "size": nested_name(creature.get("size")),
        "type": nested_name(creature.get("type")),
        "subtype": "",
        "tags": [],
        "alignment": clean(creature.get("alignment")),
        "ac": creature.get("armor_class"),
        "hp": creature.get("hit_points"),
        "hit_dice": clean(creature.get("hit_dice")),
        "speed": _speed_string(creature),
        "initiative": creature.get("initiative_bonus"),
        "str": _score("strength"), "dex": _score("dexterity"), "con": _score("constitution"),
        "int": _score("intelligence"), "wis": _score("wisdom"), "cha": _score("charisma"),
        "cr": cr_display(creature.get("challenge_rating")),
        "xp": creature.get("experience_points"),
        "proficiency_bonus": creature.get("proficiency_bonus"),
        "senses": senses_display(creature),
        "languages": clean(languages.get("as_string") if isinstance(languages, dict) else languages),
        "damage_vulnerabilities": _split(defenses.get("damage_vulnerabilities_display")),
        "damage_resistances": _split(defenses.get("damage_resistances_display")),
        "damage_immunities": _split(defenses.get("damage_immunities_display")),
        "condition_immunities": _split(defenses.get("condition_immunities_display")),
        "special_abilities": _named_entries(creature.get("traits") or []),
        "actions": _named_entries(actions_of_type(creature, "ACTION")),
        "bonus_actions": _named_entries(actions_of_type(creature, "BONUS_ACTION")),
        "reactions": _named_entries(actions_of_type(creature, "REACTION")),
        "legendary_actions": _named_entries(actions_of_type(creature, "LEGENDARY_ACTION")),
        # Provenance so the app can attribute/license correctly and dedupe if needed.
        "source_key": clean(document.get("key")),
        "source_edition": (nested_name(document.get("gamesystem") or {}) or "").strip(),
    }


def write_app_json(creatures: List[Dict[str, Any]], out_dir: str) -> None:
    """Write per-edition monster files in the app's shape. 5e-2024 -> monsters-2024.json,
    everything else -> monsters-2014.json (most OGL third-party content is 2014-compatible)."""
    os.makedirs(out_dir, exist_ok=True)
    buckets: Dict[str, List[Dict[str, Any]]] = {"2024": [], "2014": []}
    for creature in creatures:
        gs = ((creature.get("document") or {}).get("gamesystem") or {}).get("key") or ""
        edition = "2024" if gs == "5e-2024" else "2014"
        buckets[edition].append(app_monster(creature))

    for edition, rows in buckets.items():
        rows.sort(key=lambda m: m.get("name") or "")
        path = os.path.join(out_dir, f"monsters-{edition}.json")
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(rows, handle, ensure_ascii=False, indent=2)
        print(f"  wrote {path}: {len(rows)} monsters")


# ---------------------------------------------------------------- cli


def main() -> int:
    parser = argparse.ArgumentParser(description="Export Open5e creature stat blocks to xlsx.")
    parser.add_argument("--out", default="", help="xlsx workbook path (omit to skip the workbook)")
    parser.add_argument("--json", default="", help="directory to write the app's monsters-2024.json and monsters-2014.json")
    parser.add_argument("--documents", default="", help="comma-separated source keys, e.g. srd-2014,srd-2024")
    parser.add_argument("--gamesystem", default="", help="filter by game system key, e.g. 5e-2014 or 5e-2024")
    parser.add_argument("--cache", default="", help="read from / write to this raw JSON file")
    parser.add_argument("--list-documents", action="store_true", help="print source keys and counts, then exit")
    args = parser.parse_args()

    creatures = load_or_fetch(args.cache or None)
    print(f"Retrieved {len(creatures)} creatures")

    if args.list_documents:
        print(f"\n{'key':<14} {'creatures':>9}  document")
        for row in source_rows(creatures):
            print(f"{row[0]:<14} {row[4]:>9}  {row[1]} ({row[2]})")
        return 0

    if args.documents:
        wanted = {key.strip() for key in args.documents.split(",") if key.strip()}
        creatures = [c for c in creatures if (c.get("document") or {}).get("key") in wanted]
        print(f"Filtered to documents {sorted(wanted)}: {len(creatures)} creatures")

    if args.gamesystem:
        target = args.gamesystem.strip()
        creatures = [
            c
            for c in creatures
            if ((c.get("document") or {}).get("gamesystem") or {}).get("key") == target
        ]
        print(f"Filtered to game system {target}: {len(creatures)} creatures")

    if not creatures:
        print("Nothing matched those filters, so no workbook was written.", file=sys.stderr)
        return 1

    creatures.sort(key=lambda c: ((c.get("document") or {}).get("key") or "", c.get("name") or ""))

    # Default when neither output is named: write the app JSON to the current directory, since that's
    # the point of the tool for Six Axes. Pass --out to also (or instead) write the xlsx workbook.
    wrote_something = False
    if args.json or not args.out:
        target = args.json or "."
        print(f"Writing app JSON to {target}")
        write_app_json(creatures, target)
        wrote_something = True
    if args.out:
        print(f"Writing {args.out}")
        write_workbook(creatures, args.out)
        wrote_something = True

    if not wrote_something:
        print("No output written.", file=sys.stderr)
        return 1
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
